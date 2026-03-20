#!/usr/bin/env python3
"""End-to-end test for template fill functionality"""
import urllib.request
import json
import sys
import os
import time

BASE_URL = "http://localhost:18080/api/v1"
TOKEN = None

def get_token():
    global TOKEN
    # Login as drw
    data = json.dumps({"username": "drw", "password": "drw"}).encode()
    req = urllib.request.Request(f"{BASE_URL}/users/auth", data=data, headers={"Content-Type": "application/json"}, method="POST")
    resp = json.loads(urllib.request.urlopen(req, timeout=10).read().decode())
    TOKEN = resp["data"]["token"]
    print(f"Login OK (userId={resp['data']['userId']})")

def api_json(method, path, data=None):
    headers = {"Authorization": f"Bearer {TOKEN}"}
    if data is not None:
        body = json.dumps(data).encode()
        headers["Content-Type"] = "application/json"
        req = urllib.request.Request(f"{BASE_URL}{path}", data=body, headers=headers, method=method)
    else:
        req = urllib.request.Request(f"{BASE_URL}{path}", headers=headers, method=method)
    try:
        resp = urllib.request.urlopen(req, timeout=120)
        return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f"  HTTP {e.code}: {body[:300]}")
        return None

def api_upload(path, file_path, field_name="file"):
    boundary = "----FormBoundary" + str(int(time.time()))
    filename = os.path.basename(file_path)
    with open(file_path, "rb") as f:
        file_data = f.read()
    
    parts = []
    parts.append(f"--{boundary}".encode())
    parts.append(f'Content-Disposition: form-data; name="{field_name}"; filename="{filename}"'.encode())
    parts.append(b"Content-Type: application/octet-stream")
    parts.append(b"")
    parts.append(file_data)
    parts.append(f"--{boundary}--".encode())
    body = b"\r\n".join(parts)
    
    headers = {
        "Content-Type": f"multipart/form-data; boundary={boundary}",
        "Authorization": f"Bearer {TOKEN}"
    }
    req = urllib.request.Request(f"{BASE_URL}{path}", data=body, headers=headers, method="POST")
    try:
        resp = urllib.request.urlopen(req, timeout=60)
        return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f"  HTTP {e.code}: {body[:300]}")
        return None

def main():
    print("=== DocAI Fill Test ===\n")
    
    get_token()
    
    # Check existing source documents
    print("\n[1] Check source documents...")
    resp = api_json("GET", "/source/documents")
    docs = resp.get("data", []) if resp else []
    parsed_docs = [d for d in docs if d.get("uploadStatus") == "parsed"]
    print(f"  Total docs: {len(docs)}, Parsed: {len(parsed_docs)}")
    for d in parsed_docs:
        print(f"  id={d['id']}, name={d['fileName']}")
    
    if not parsed_docs:
        # Upload source
        print("\n[1b] Uploading source document...")
        src_path = r"f:\DocAI\docai-pro\data\local-oss\source_documents\中国COVID-19新冠疫情情况.docx"
        resp = api_upload("/source/upload", src_path)
        if not resp or resp.get("code") != 200:
            print(f"  Upload failed: {resp}")
            return
        doc_id = resp["data"]["id"]
        print(f"  Uploaded docId={doc_id}, waiting for extraction...")
        for i in range(30):
            time.sleep(3)
            sr = api_json("GET", f"/source/{doc_id}")
            if sr and sr.get("data"):
                st = sr["data"].get("uploadStatus", "")
                print(f"  [{i*3}s] Status: {st}")
                if st == "parsed":
                    parsed_docs = [{"id": doc_id}]
                    break
    
    # Upload template
    print("\n[2] Upload template...")
    tpl_path = r"f:\DocAI\docai-pro\data\local-oss\template_files\COVID-19 模板.xlsx"
    resp = api_upload("/template/upload", tpl_path)
    if not resp or resp.get("code") != 200:
        print(f"  Upload failed: {resp}")
        return
    tpl_id = resp["data"]["id"]
    print(f"  Template uploaded: id={tpl_id}")
    
    # Parse slots
    print("\n[3] Parse slots...")
    resp = api_json("POST", f"/template/{tpl_id}/parse")
    if resp:
        slots = resp.get("data", [])
        print(f"  Slots: {len(slots)}")
        for s in slots:
            print(f"    {s['label']} - {s.get('slotType', '')} @ {s['position']}")
    
    # Fill
    print("\n[4] Fill template...")
    # Only use COVID-related docs, not all parsed docs
    covid_docs = [d for d in parsed_docs if "COVID" in d.get("fileName", "") or "新冠" in d.get("fileName", "")]
    if not covid_docs:
        covid_docs = parsed_docs
    doc_ids = [d["id"] for d in covid_docs]
    print(f"  Using {len(doc_ids)} source docs: {doc_ids}")
    start = time.time()
    resp = api_json("POST", f"/template/{tpl_id}/fill", {"docIds": doc_ids, "userRequirement": ""})
    elapsed = time.time() - start
    print(f"  Completed in {elapsed:.1f}s")
    if resp:
        print(f"  Result: {json.dumps(resp, ensure_ascii=False)[:500]}")
        output = resp.get("data", {}).get("outputFile") if resp.get("data") else None
        if output:
            print(f"\n  Output: {output}")
            # Read and display result
            import openpyxl
            wb = openpyxl.load_workbook(output)
            ws = wb.active
            print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
            for row in ws.iter_rows(min_row=1, max_row=min(35, ws.max_row), values_only=True):
                vals = [str(v)[:20] if v else "" for v in row]
                if any(v for v in vals):
                    print(f"  | {'  | '.join(vals)} |")
    
    print("\n=== Test Complete ===")

if __name__ == "__main__":
    main()
