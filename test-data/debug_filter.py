#!/usr/bin/env python3
"""Debug: what source values match paragraph text"""
import openpyxl

wb = openpyxl.load_workbook(
    r'F:\DocAI\docai-pro\data\local-oss\source_documents\山东省环境空气质量监测数据信息202512171921_0.xlsx',
    data_only=True
)
ws = wb.active
hdr = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print(f'Sheet: {ws.title}')
print(f'Header: {hdr}')

col_vals = {}
for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=501, values_only=True)):
    for c_idx, val in enumerate(row):
        if val is not None:
            s = str(val).strip()
            if s and len(s) <= 30:
                col_vals.setdefault(c_idx, set()).add(s)

print(f'\nCol unique counts:')
for c, v in col_vals.items():
    name = hdr[c] if c < len(hdr) else f'col{c}'
    print(f'  Col {c} ({name}): {len(v)} unique values')
    if len(v) <= 20:
        print(f'    Values: {sorted(v)}')

# Test paragraph for table 0 (德州)
para0 = '本表记录2025年11月25日09:00时刻德州市各监测站点的空气质量检测数据。涵盖城市、行政区、站点名称、AQI指数、PM10、PM2.5浓度、首要污染物及污染类型等核心指标，实时反映德州市大气环境质量状况'
para1 = '本表记录2025年11月25日09:00时刻潍坊市各监测站点的空气质量检测数据。结构与德州市一致，便于城市间数据对比分析。'
para2 = '本表记录2025年11月25日09:00时刻临沂市各监测站点的空气质量检测数据。用于系统记录临沂市环境空气质量监测结果'
user_req = '完成填表工作，要求提取表格中对应数据'

for ti, (para, name) in enumerate([(para0, '德州'), (para1, '潍坊'), (para2, '临沂')]):
    context = para + '\n' + user_req
    print(f'\n=== Table {ti} ({name}) ===')
    best_col = -1
    best_val = None
    best_spec = 0
    for c_idx, vals in col_vals.items():
        if len(vals) <= 1 or len(vals) > 100:
            continue
        for v in vals:
            if len(v) < 2:
                continue
            if v in context and len(v) > best_spec:
                best_col = c_idx
                best_val = v
                best_spec = len(v)
                print(f'  Match: Col {c_idx} ({hdr[c_idx] if c_idx < len(hdr) else "?"}), value="{v}" (len={len(v)})')

    print(f'  BEST: Col {best_col}, value="{best_val}" (len={best_spec})')

wb.close()
