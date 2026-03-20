#!/usr/bin/env python3
"""Test multi-source merge: xlsx + docx source documents filling a single template"""
import urllib.request
import json
import os
import time

BASE_URL = "http://localhost:18080/api/v1"
TOKEN = None

def get_token():
    global TOKEN
    data = json.dumps({"username": "drw", "password": "drw"}).encode()
    req = urllib.request.Request(f"{BASE_URL}/users/auth", data=data, headers={"Content-Type": "application/json"}, method="POST")
    resp = json.loads(urllib.request.urlopen(req, timeout=10).read().decode())
    TOKEN = resp["data"]["token"]
    print(f"Login OK (userId={resp['data']['userId']})")

def api_json(method, path, data=None):
    headers = {"Authorization": f"Bearer {TOKEN}"}
    if data is not None:
        body = json.dumps(data).encode()
        headers["Content-Type"] = "application/json"
        req = urllib.request.Request(f"{BASE_URL}{path}", data=body, headers=headers, method=method)
    else:
        req = urllib.request.Request(f"{BASE_URL}{path}", headers=headers, method=method)
    try:
        resp = urllib.request.urlopen(req, timeout=120)
        return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f"  HTTP {e.code}: {body[:300]}")
        return None

def api_upload(path, file_path, field_name="file"):
    boundary = "----FormBoundary" + str(int(time.time()))
    filename = os.path.basename(file_path)
    with open(file_path, "rb") as f:
        file_data = f.read()
    parts = []
    parts.append(f"--{boundary}".encode())
    parts.append(f'Content-Disposition: form-data; name="{field_name}"; filename="{filename}"'.encode())
    parts.append(b"Content-Type: application/octet-stream")
    parts.append(b"")
    parts.append(file_data)
    parts.append(f"--{boundary}--".encode())
    body = b"\r\n".join(parts)
    headers = {
        "Content-Type": f"multipart/form-data; boundary={boundary}",
        "Authorization": f"Bearer {TOKEN}"
    }
    req = urllib.request.Request(f"{BASE_URL}{path}", data=body, headers=headers, method="POST")
    try:
        resp = urllib.request.urlopen(req, timeout=60)
        return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f"  HTTP {e.code}: {body[:300]}")
        return None

def wait_extraction(doc_id, max_wait=120):
    for i in range(max_wait // 3):
        time.sleep(3)
        sr = api_json("GET", "/source/documents")
        if sr:
            for d in sr.get("data", []):
                if d["id"] == doc_id:
                    st = d.get("uploadStatus", "")
                    print(f"    [{i*3}s] Status: {st}")
                    if st == "parsed":
                        return d
    return None

def main():
    print("=== Multi-Source Merge Test (xlsx + docx) ===\n")
    get_token()

    # Upload xlsx source document
    print("\n[1] Upload xlsx source document...")
    xlsx_path = r"f:\DocAI\docai-pro\data\local-oss\source_documents\COVID-19_extra.xlsx"
    resp = api_upload("/source/upload", xlsx_path)
    if not resp or resp.get("code") != 200:
        print(f"  Upload failed: {resp}")
        return
    xlsx_doc_id = resp["data"]["id"]
    print(f"  xlsx docId={xlsx_doc_id}, waiting for extraction...")
    xlsx_doc = wait_extraction(xlsx_doc_id)
    if not xlsx_doc:
        print("  xlsx extraction timeout!")
        return
    print(f"  xlsx extraction done")

    # Upload docx source document
    print("\n[2] Upload docx source document...")
    docx_path = r"f:\DocAI\docai-pro\data\local-oss\source_documents\中国COVID-19新冠疫情情况.docx"
    resp = api_upload("/source/upload", docx_path)
    if not resp or resp.get("code") != 200:
        print(f"  Upload failed: {resp}")
        return
    docx_doc_id = resp["data"]["id"]
    print(f"  docx docId={docx_doc_id}, waiting for extraction...")
    docx_doc = wait_extraction(docx_doc_id)
    if not docx_doc:
        print("  docx extraction timeout!")
        return
    print(f"  docx extraction done")

    # Upload template
    print("\n[3] Upload template...")
    tpl_path = r"f:\DocAI\docai-pro\data\local-oss\template_files\COVID-19 模板.xlsx"
    resp = api_upload("/template/upload", tpl_path)
    if not resp or resp.get("code") != 200:
        print(f"  Upload failed: {resp}")
        return
    tpl_id = resp["data"]["id"]
    print(f"  Template id={tpl_id}")

    # Parse slots
    print("\n[4] Parse slots...")
    resp = api_json("POST", f"/template/{tpl_id}/parse")
    if resp:
        slots = resp.get("data", [])
        print(f"  Slots: {len(slots)}")
        for s in slots:
            print(f"    {s['label']} - {s.get('slotType', '')} @ {s['position']}")

    # Fill with BOTH source docs
    print("\n[5] Fill with both xlsx + docx sources...")
    doc_ids = [xlsx_doc_id, docx_doc_id]
    print(f"  docIds: {doc_ids}")
    start = time.time()
    resp = api_json("POST", f"/template/{tpl_id}/fill", {"docIds": doc_ids, "userRequirement": ""})
    elapsed = time.time() - start
    print(f"  Completed in {elapsed:.1f}s")
    
    if resp:
        output = resp.get("data", {}).get("outputFile") if resp.get("data") else None
        if output:
            print(f"\n  Output: {output}")
            import openpyxl
            wb = openpyxl.load_workbook(output)
            ws = wb.active
            print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
            
            # Check for xlsx provinces (浙江省, 山东省, 陕西省)
            xlsx_provinces = set()
            docx_provinces = set()
            all_provinces = []
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                vals = [str(v)[:25] if v else "" for v in row]
                if any(v for v in vals):
                    print(f"  | {'  | '.join(vals)} |")
                    name = vals[0] if vals else ""
                    if name in ("浙江省", "山东省", "陕西省"):
                        xlsx_provinces.add(name)
                    elif name and name != "国家/地区":
                        docx_provinces.add(name)
                    if name and name != "国家/地区":
                        all_provinces.append(name)
            
            print(f"\n  === Merge Summary ===")
            print(f"  Total data rows: {len(all_provinces)}")
            print(f"  From xlsx: {xlsx_provinces}")
            print(f"  From docx: {len(docx_provinces)} provinces")
            if xlsx_provinces and docx_provinces:
                print(f"  MERGE SUCCESS: Both xlsx and docx data present!")
            elif xlsx_provinces:
                print(f"  MERGE FAILED: Only xlsx data present")
            elif docx_provinces:
                print(f"  MERGE FAILED: Only docx data present")
            else:
                print(f"  MERGE FAILED: No data")
    
    print("\n=== Test Complete ===")

if __name__ == "__main__":
    main()
